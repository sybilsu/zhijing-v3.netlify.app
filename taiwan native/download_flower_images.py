"""
下載「花」欄位圖片，存為「花-植物名.jpg」
"""
import sys, io, re, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl, requests
from pathlib import Path
from urllib.parse import urlparse, unquote

EXCEL   = r'C:\Users\g1375\web app\台灣中部原生植物資料庫 500種\thesis database\2026_植物屬性表_500種.xlsx'
OUT_DIR = Path(r'C:\Users\g1375\web app\台灣中部原生植物資料庫 500種\thesis database')

HEADERS = {
    'User-Agent': 'PlantDatabaseBot/1.0 (educational research; contact chuan.sybil@gmail.com)'
}
BASE_DELAY  = 2.0
MAX_RETRIES = 4

def get_ext(url: str) -> str:
    ext = Path(unquote(urlparse(url).path)).suffix.lower()
    return ext if ext in ('.jpg', '.png', '.gif', '.webp') else ('.jpg' if ext != '.jpeg' else '.jpg')

def safe_name(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', s.strip())

def download(url: str, out: Path) -> bool:
    if out.exists():
        print(f'  [skip] {out.name}')
        return True
    for attempt in range(MAX_RETRIES + 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=30, stream=True)
            if r.status_code == 429:
                wait = BASE_DELAY * (3 ** attempt)
                print(f'  [429]  {out.name} — 等 {wait:.0f}s ({attempt+1}/{MAX_RETRIES})')
                time.sleep(wait)
                continue
            r.raise_for_status()
            with open(out, 'wb') as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
            print(f'  [ok]   {out.name}')
            return True
        except requests.exceptions.HTTPError:
            if attempt < MAX_RETRIES:
                continue
            print(f'  [err]  {out.name}: HTTP error after {MAX_RETRIES} retries')
            return False
        except Exception as e:
            print(f'  [err]  {out.name}: {e}')
            return False
    return False

def main():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    ok = fail = 0
    total = ws.max_row - 1

    for r in range(2, ws.max_row + 1):
        col5 = ws.cell(r, 5).value   # 花 URL
        col6 = ws.cell(r, 6).value   # 植物名

        if not col6:
            continue
        plant = safe_name(col6.split('/')[0].strip())

        if col5 and str(col5).startswith('http'):
            url = col5.strip()
            out = OUT_DIR / f'花-{plant}{get_ext(url)}'
            if download(url, out):
                ok += 1
            else:
                fail += 1
            time.sleep(BASE_DELAY)

        if (r - 1) % 50 == 0:
            print(f'--- 進度 {r-1}/{total}，成功 {ok}，失敗 {fail} ---')

    print(f'\n完成：下載成功 {ok}，失敗 {fail}')

if __name__ == '__main__':
    main()
