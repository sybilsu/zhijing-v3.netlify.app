"""
下載植物圖片並依「葉/近/全株-植物名」命名
"""
import sys, io, os, re, time, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import requests
from pathlib import Path
from urllib.parse import urlparse, unquote

EXCEL = r'C:\Users\g1375\web app\台灣中部原生植物資料庫 500種\thesis database\2026_植物屬性表_500種.xlsx'
OUT_DIR = Path(r'C:\Users\g1375\web app\台灣中部原生植物資料庫 500種\thesis database')

HEADERS = {
    'User-Agent': 'PlantDatabaseBot/1.0 (educational research; contact chuan.sybil@gmail.com)'
}

BASE_DELAY = 2.0   # 秒，每次請求基本間隔
MAX_RETRIES = 4    # 429 時最多重試次數

def get_ext(url: str, content_type: str = '') -> str:
    path = unquote(urlparse(url).path)
    ext = Path(path).suffix.lower()
    if ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
        return ext if ext != '.jpeg' else '.jpg'
    if 'png' in content_type:
        return '.png'
    if 'gif' in content_type:
        return '.gif'
    if 'webp' in content_type:
        return '.webp'
    return '.jpg'

def safe_name(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', name.strip())

def download(url: str, out_path: Path) -> bool:
    if out_path.exists():
        print(f'  [skip] 已存在: {out_path.name}')
        return True
    for attempt in range(MAX_RETRIES + 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=30, stream=True)
            if r.status_code == 429:
                wait = BASE_DELAY * (3 ** attempt)  # 指數退避: 2, 6, 18, 54 秒
                print(f'  [429]  {out_path.name} — 等待 {wait:.0f}s 後重試 ({attempt+1}/{MAX_RETRIES})')
                time.sleep(wait)
                continue
            r.raise_for_status()
            with open(out_path, 'wb') as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
            print(f'  [ok]   {out_path.name}')
            return True
        except requests.exceptions.HTTPError:
            if attempt < MAX_RETRIES:
                continue
            print(f'  [err]  {out_path.name}: HTTP error after {MAX_RETRIES} retries')
            return False
        except Exception as e:
            print(f'  [err]  {out_path.name}: {e}')
            return False
    return False

def fetch_url(url: str, prefix: str, plant_name: str) -> bool:
    ext = get_ext(url)
    out = OUT_DIR / f'{prefix}-{plant_name}{ext}'
    result = download(url, out)
    time.sleep(BASE_DELAY)
    return result

def main():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active

    ok = fail = skip_exist = renamed = 0
    total = ws.max_row - 1
    done = 0

    for r in range(2, ws.max_row + 1):
        col2 = ws.cell(r, 2).value
        col3 = ws.cell(r, 3).value
        col4 = ws.cell(r, 4).value
        col5 = ws.cell(r, 5).value

        if not col5:
            continue

        plant_name = safe_name(col5.split('/')[0].strip())
        done += 1

        if col2 and col2.startswith('http'):
            if fetch_url(col2.strip(), '葉', plant_name):
                ok += 1
            else:
                fail += 1

        if col3 and col3.startswith('http'):
            if fetch_url(col3.strip(), '近', plant_name):
                ok += 1
            else:
                fail += 1

        if col4:
            src = OUT_DIR / col4
            dst = OUT_DIR / f'全株-{plant_name}.jpg'
            if src.exists() and not dst.exists():
                shutil.copy2(src, dst)
                print(f'  [rename] {col4} → {dst.name}')
                renamed += 1
            elif dst.exists():
                skip_exist += 1

        if done % 20 == 0:
            print(f'--- 進度 {done}/{total}，成功 {ok}，失敗 {fail} ---')

    print(f'\n完成：下載成功 {ok}，失敗 {fail}，全株改名 {renamed}，已存在跳過 {skip_exist}')

if __name__ == '__main__':
    main()
