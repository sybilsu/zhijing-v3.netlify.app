"""
為缺少圖片的植物從 Wikipedia / Wikimedia Commons 搜尋並下載影像
類別：葉、近、花、全株
"""
import sys, io, re, time, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl, requests
from pathlib import Path
from urllib.parse import quote, urlparse, unquote

EXCEL   = r'C:\Users\g1375\web app\台灣中部原生植物資料庫 500種\thesis database\2026_植物屬性表_500種.xlsx'
OUT_DIR = Path(r'C:\Users\g1375\web app\台灣中部原生植物資料庫 500種\thesis database')
LOG     = OUT_DIR / 'missing_images_log.txt'

HEADERS = {'User-Agent': 'PlantDatabaseBot/1.0 (educational research; contact chuan.sybil@gmail.com)'}
DELAY       = 1.5   # 基本請求間隔（秒）
MAX_RETRIES = 3

EXTS = ('.jpg', '.png', '.gif', '.webp')

# 各類別的 Wikimedia 搜尋提示詞
HINTS = {
    '葉':  'leaf foliage',
    '近':  'close up',
    '花':  'flower bloom',
    '全株': 'plant habit',
}

def safe_name(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', s.strip())

def binomial(sci: str) -> str:
    parts = sci.strip().split()
    return f'{parts[0]} {parts[1]}' if len(parts) >= 2 else sci.strip()

def already_exists(prefix: str, plant: str) -> bool:
    return any((OUT_DIR / f'{prefix}-{plant}{e}').exists() for e in EXTS)

def get_ext_from_url(url: str) -> str:
    ext = Path(unquote(urlparse(url).path)).suffix.lower()
    return ext if ext in EXTS else '.jpg'

# ── Wikipedia REST API ──────────────────────────────────────────────────────
def wiki_page_image(title: str) -> str | None:
    url = f'https://en.wikipedia.org/api/rest_v1/page/summary/{quote(title)}'
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            d = r.json()
            src = (d.get('originalimage') or d.get('thumbnail') or {}).get('source')
            return src
    except Exception:
        pass
    return None

# ── Wikimedia Commons 搜尋 ──────────────────────────────────────────────────
def commons_search(query: str) -> str | None:
    try:
        r = requests.get('https://commons.wikimedia.org/w/api.php', params={
            'action': 'query', 'list': 'search',
            'srsearch': query, 'srnamespace': 6,
            'format': 'json', 'srlimit': 5,
        }, headers=HEADERS, timeout=15)
        hits = r.json().get('query', {}).get('search', [])
        if not hits:
            return None
        title = hits[0]['title']
        r2 = requests.get('https://commons.wikimedia.org/w/api.php', params={
            'action': 'query', 'titles': title,
            'prop': 'imageinfo', 'iiprop': 'url',
            'iiurlwidth': 500, 'format': 'json',
        }, headers=HEADERS, timeout=15)
        for p in r2.json().get('query', {}).get('pages', {}).values():
            ii = p.get('imageinfo', [])
            if ii:
                return ii[0].get('thumburl') or ii[0].get('url')
    except Exception:
        pass
    return None

# ── 下載 ────────────────────────────────────────────────────────────────────
def download(url: str, out: Path) -> bool:
    for attempt in range(MAX_RETRIES + 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=30, stream=True)
            if r.status_code == 429:
                wait = DELAY * (3 ** attempt)
                time.sleep(wait)
                continue
            r.raise_for_status()
            with open(out, 'wb') as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
            return True
        except Exception:
            if attempt >= MAX_RETRIES:
                return False
    return False

# ── 搜尋並下載單一類別 ───────────────────────────────────────────────────────
def find_image(plant: str, sci: str, prefix: str) -> tuple[bool, str]:
    """回傳 (成功, 狀態)"""
    out_base = OUT_DIR / f'{prefix}-{plant}'
    # 已存在則跳過
    for e in EXTS:
        if (OUT_DIR / f'{prefix}-{plant}{e}').exists():
            return True, 'skip'

    hint  = HINTS.get(prefix, '')
    binom = binomial(sci) if sci else plant

    # 嘗試順序：
    # 1. Wikipedia 頁面縮圖（依學名）
    img = wiki_page_image(binom)
    time.sleep(DELAY)

    # 2. Wikimedia Commons 搜尋（學名 + 類別提示）
    if not img:
        img = commons_search(f'{binom} {hint}')
        time.sleep(DELAY)

    # 3. Wikimedia Commons 搜尋（中文名）
    if not img:
        img = commons_search(f'{plant} {hint}')
        time.sleep(DELAY)

    # 4. Wikipedia 搜尋（中文維基）
    if not img:
        img = wiki_page_image(plant)
        time.sleep(DELAY)

    if not img:
        return False, 'not_found'

    ext = get_ext_from_url(img)
    out = OUT_DIR / f'{prefix}-{plant}{ext}'
    if download(img, out):
        return True, 'found'
    return False, 'download_err'

# ── 主程式 ──────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active

    # 欄位對應: prefix → Excel col
    COL_MAP = {'葉': 2, '近': 3, '花': 5}  # 全株(col4) 僅為本地檔名，無 URL，統一走搜尋

    ok = found = fail = skipped = not_found = 0
    log_lines = []
    total_rows = ws.max_row - 1
    done = 0

    for r in range(2, ws.max_row + 1):
        c6 = ws.cell(r, 6).value
        if not c6:
            continue
        plant = safe_name(c6.split('/')[0].strip())
        sci   = c6.split('/')[-1].strip() if '/' in c6 else ''
        done += 1

        for prefix, col in COL_MAP.items():
            cell_val = ws.cell(r, col).value
            # 若 Excel 有 URL 且檔案已在本次之前的下載腳本取得，skip
            if already_exists(prefix, plant):
                skipped += 1
                continue
            # Excel 已有 URL 但未下載成功（重試）
            if cell_val and str(cell_val).startswith('http'):
                out_base = OUT_DIR / f'{prefix}-{plant}'
                ext = get_ext_from_url(str(cell_val))
                out = OUT_DIR / f'{prefix}-{plant}{ext}'
                if download(str(cell_val).strip(), out):
                    print(f'  [retry-ok] {out.name}')
                    ok += 1
                    time.sleep(DELAY)
                    continue
            # 從網路搜尋
            success, status = find_image(plant, sci, prefix)
            if status == 'skip':
                skipped += 1
            elif status == 'found':
                print(f'  [web-ok]  {prefix}-{plant}')
                found += 1
            elif status == 'not_found':
                print(f'  [miss]    {prefix}-{plant}')
                log_lines.append(f'not_found: {prefix}-{plant} | {sci}')
                not_found += 1
            else:
                print(f'  [err]     {prefix}-{plant}: {status}')
                log_lines.append(f'err: {prefix}-{plant} | {sci}')
                fail += 1

        # 全株 — 無論 col4 有無本地檔名，只要目的地不存在就搜尋
        if not already_exists('全株', plant):
            success, status = find_image(plant, sci, '全株')
            if status == 'skip':
                skipped += 1
            elif status == 'found':
                print(f'  [web-ok]  全株-{plant}')
                found += 1
            elif status == 'not_found':
                log_lines.append(f'not_found: 全株-{plant} | {sci}')
                not_found += 1
            else:
                log_lines.append(f'err: 全株-{plant} | {sci}')
                fail += 1
        else:
            skipped += 1

        if done % 25 == 0:
            print(f'=== 進度 {done}/{total_rows} | 已找到 {found} | 跳過 {skipped} | 找不到 {not_found} | 錯誤 {fail} ===')

    # 寫 log
    with open(LOG, 'w', encoding='utf-8') as f:
        f.write('\n'.join(log_lines))

    print(f'\n完成：從網路找到 {found}，已存在跳過 {skipped}，找不到 {not_found}，下載錯誤 {fail}')
    print(f'找不到的植物名單：{LOG}')

if __name__ == '__main__':
    main()
